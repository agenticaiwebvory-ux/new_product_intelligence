"""
Product Scraper — Main Runner (v4)
Now returns an Excel file path + updates a progress dict for the API.

Usage (CLI):
    python run.py --input <pdf_or_csv> --brand <brand_name>
"""
import argparse
import csv
import os
import re
import sys
from datetime import datetime
from config import CACHE_COLUMNS, OUTPUT_DIR

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
from cache import get_cached_styles, get_cached_results_for_vendor, append_to_cache
from serper_search import batch_search, _allowlist_pass
from filter import filter_results
from pdf_extractor import extract_styles_from_pdf
from csv_reader import read_styles_from_csv

CSV_COLUMNS = [
    "Source PDF", "Style No", "Description", "Color", "Sizes Available",
    "Total Qty", "Inventory Detail", "Wholesale Price", "Retail Price",
    "Image Link", "Season", "Fabric", "Country of Origin", "Vendor"
]

def run(input_path: str, brand_name: str, force: bool = False, progress: dict = None):
    """
    Main scraper entry point.

    Args:
        input_path: Path to PDF or CSV file
        brand_name: Vendor/brand name
        force: If True, re-scrape even if style is in cache
        progress: Optional dict updated in-place for API progress tracking

    Returns:
        str: Path to the output Excel file
    """
    if progress is None:
        progress = {}

    def update(step, detail="", pct=0):
        progress["step"] = step
        progress["detail"] = detail
        progress["percent"] = pct
        print(f"  [{pct}%] {step}: {detail}")

    print(f"\n{'='*60}")
    print(f"  Product Scraper — {brand_name}")
    print(f"  Input: {os.path.basename(input_path)}")
    print(f"{'='*60}\n")

    # ── Step 1: Extract style numbers ──
    update("extracting", "Reading input file...", 5)
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".pdf":
        styles = extract_styles_from_pdf(input_path)
    elif ext in (".csv", ".tsv", ".xlsx"):
        styles = read_styles_from_csv(input_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    if not styles:
        raise ValueError("No style numbers found in input file!")

    all_style_numbers = {(s["Style No"] or "").strip().upper() for s in styles}
    update("extracting", f"Found {len(all_style_numbers)} unique styles", 10)
    progress["total_styles"] = len(all_style_numbers)

    # ── Step 2: Check cache ──
    update("cache_check", "Checking cache...", 15)
    cached_styles = get_cached_styles(vendor_name=brand_name)
    cached_results = get_cached_results_for_vendor(brand_name)

    already_done = all_style_numbers & cached_styles
    needs_search = all_style_numbers - cached_styles

    if force:
        needs_search = all_style_numbers
        already_done = set()

    progress["cached"] = len(already_done)
    progress["needs_search"] = len(needs_search)
    update("cache_check", f"{len(already_done)} cached, {len(needs_search)} new", 20)

    # ── Step 3: Serper search (parallel — up to 20 concurrent) ──
    new_results = []
    if needs_search:
        styles_to_search = [s for s in styles if s["Style No"] in needs_search]
        total_to_search = len(styles_to_search)

        def _search_progress(completed, total, style_no):
            pct = 20 + int((completed / total) * 55)  # 20% → 75%
            update("searching", f"[{completed}/{total}] {style_no}", pct)

        update("searching", f"Running parallel Serper for {total_to_search} styles (20 workers)...", 25)
        raw_results = batch_search(styles_to_search, brand_name, progress_callback=_search_progress)

        update("filtering", "Filtering results...", 75)
        new_results = filter_results(raw_results, brand_name)

        # Save to cache (including styles with no results)
        cache_rows = []
        styles_with_results = {r.get("Style No") for r in new_results}
        
        # Add actual results
        for r in new_results:
            cache_rows.append({
                "Style No":         r.get("Style No", ""),
                "Description":      r.get("Description", ""),
                "Colors":           r.get("Colors", ""),
                "Sizes":            r.get("Sizes", ""),
                "Total Qty":        r.get("Total Qty", ""),
                "Inventory Detail": r.get("Inventory Detail", ""),
                "Product URL":      r.get("Product URL", ""),
                "Price (USD)":      r.get("Price (USD)", ""),
                "Image Link":       r.get("Image Link", ""),
                "source_title":     r.get("source_title", ""),
                "Vendor":           brand_name,
            })
            
        # Add "negative" results for styles that yielded nothing
        for sno in needs_search:
            if sno not in styles_with_results:
                cache_rows.append({
                    "Style No":         sno,
                    "Product URL":      "",
                    "Vendor":           brand_name,
                })
                
        if cache_rows:
            append_to_cache(cache_rows)
    else:
        update("searching", "All styles in cache — no Serper calls needed", 75)

    # ── Step 4: Combine results ──
    update("building_output", "Building Excel...", 80)
    
    # Map results by Style No
    results_by_style = {}
    for r in cached_results:
        url = r.get("Product URL", "")
        # If it's in the cache, we trust it already passed the allowlist
        
        # Drop duplicates across cache
        sno_key = str(r.get("Style No", "")).strip().upper()
        if url and any(x.get("Product URL") == url for x in results_by_style.get(sno_key, [])):
            continue
        results_by_style.setdefault(sno_key, []).append(r)
        
    for r in new_results:
        url = r.get("Product URL", "")
        sno_key = str(r.get("Style No", "")).strip().upper()
        if url and any(x.get("Product URL") == url for x in results_by_style.get(sno_key, [])):
            continue
        results_by_style.setdefault(sno_key, []).append(r)

    # Build the final flat list, preserving multiple identical styles
    combined = []
    
    for src in styles:
        sno = src.get("Style No", "")
        sno_norm = str(sno).strip().upper()
        matched_results = results_by_style.get(sno_norm, [])
        
        # We'll create one "main" row per style entry from the PDF
        main_row = {
            "Style No":          sno,
            "Description":       src.get("Description", ""),
            "Colors":            src.get("Colors", ""),
            "Sizes":             src.get("Sizes", ""),
            "Total Qty":         src.get("Total Qty", ""),
            "Inventory Detail":  src.get("Inventory Detail", ""),
            "Wholesale Price":   src.get("Wholesale Price", ""),
            "Retail Price":      src.get("Retail Price", ""),
            "Season":            src.get("Season", ""),
            "Fabric":            src.get("Fabric", ""),
            "Country of Origin": src.get("Country of Origin", ""),
            "Code":              src.get("Code", ""),
            "Vendor Name":       brand_name,
        }
        
        if not matched_results:
            main_row["Product URL"] = ""
            main_row["Image Link"] = ""
            main_row["Price (USD)"] = src.get("Retail Price") or src.get("Wholesale Price") or ""
            combined.append(main_row)
        else:
            # First result goes on the main row
            first_res = matched_results[0]
            main_row["Product URL"] = first_res.get("Product URL") or first_res.get("Product Link") or first_res.get("URL") or ""
            main_row["Image Link"] = first_res.get("Image Link") or first_res.get("Image") or ""
            main_row["Price (USD)"] = first_res.get("Price (USD)") or first_res.get("Retail Price") or src.get("Retail Price") or src.get("Wholesale Price") or ""
            combined.append(main_row)
            
            # Subsequent results go on secondary rows (metadata blanked out)
            for extra_res in matched_results[1:]:
                extra_row = {
                    "Style No":          sno, # Keep style no for reference
                    "Description":       "", "Colors": "", "Sizes": "", "Total Qty": "", 
                    "Inventory Detail":  "", "Wholesale Price": "", "Retail Price": "",
                    "Season":            "", "Fabric": "", "Country of Origin": "", 
                    "Code":              "", "Vendor Name": brand_name,
                    "Product URL":       extra_res.get("Product URL") or extra_res.get("Product Link") or extra_res.get("URL") or "",
                    "Image Link":        extra_res.get("Image Link") or extra_res.get("Image") or "",
                    "Price (USD)":       extra_res.get("Price (USD)") or extra_res.get("Retail Price") or "",
                    "is_secondary":      True # flag to hide metadata in excel
                }
                combined.append(extra_row)

    # ── Step 5: Write Output ──
    update("writing_output", "Saving CSV...", 90)
    
    # Enrich records with "Source PDF" and "Vendor" if missing
    for r in combined:
        if not r.get("Source PDF"):
            r["Source PDF"] = os.path.basename(input_path)
        if not r.get("Vendor"):
            # If pdf_extractor didn't find a specific vendor in text, use brand_name
            r["Vendor"] = brand_name

    excel_path = _write_excel(combined, brand_name, style_order=[s["Style No"] for s in styles])
    csv_path = _write_csv(combined, brand_name)
    
    unique_with_urls = len({r.get("Style No") for r in combined if r.get("Product URL")})
    total_listings = len([r for r in combined if r.get("Product URL")])

    progress["output_file"] = excel_path 
    progress["total_listings"] = total_listings
    progress["styles_with_results"] = unique_with_urls

    update("done", f"{total_listings} listings for {unique_with_urls} styles", 100)

    print(f"\n{'='*60}")
    print(f"  [OK] DONE — {brand_name}")
    print(f"  Input styles:        {len(all_style_numbers)}")
    print(f"  From cache:          {len(already_done)}")
    print(f"  New Serper calls:    {len(needs_search)}")
    print(f"  Total listings:      {total_listings}")
    print(f"  Styles with results: {unique_with_urls}")
    print(f"  Excel: {excel_path}")
    print(f"  CSV: {csv_path}")
    print(f"{'='*60}\n")

    return excel_path


def _expand_sizes(sizes_str: str) -> list:
    """
    Expand a sizes string into a list of individual size labels.

    Handles:
      '(4-16)'   → ['4','6','8','10','12','14','16']   (even steps)
      '(2-14)'   → ['2','4','6','8','10','12','14']
      '4 | 6 | 8' → ['4','6','8']
      '4, 6, 8'  → ['4','6','8']
    """
    if not sizes_str:
        return []
    sizes_str = sizes_str.strip()

    # Range format: (4-16), (2-14), (4-10) etc.
    m = re.match(r'\(?(\d+)-(\d+)\)?', sizes_str)
    if m:
        start, end = int(m.group(1)), int(m.group(2))
        step = 2 if start % 2 == 0 else 1
        return [str(s) for s in range(start, end + 1, step)]

    # List format: '4 | 6 | 8' or '4, 6, 8' or '4 6 8'
    parts = re.split(r'[|,\s]+', sizes_str)
    return [p.strip() for p in parts if p.strip().isdigit()]


def _parse_inv_detail(inv_detail: str, sizes_str: str = "") -> dict:
    """
    Parse Inventory Detail into {size_label: qty} dict.

    Handles three real-world formats found in the vendor PDFs:

    1. Standard (AP OP / Bridal):
          '4:19 | 6:32 | 8:24'
       → {4: 19, 6: 32, 8: 24}

    2. qty/size pairs (OFF PRICE):
          '2/4, 2/6'  or  '11/4, 7/6, | 11/8'
       → {4: 2, 6: 2}  /  {4: 11, 6: 7, 8: 11}

    3. Dot-separated quantities with size range/list (OFF PRICE):
          sizes_str='(4-10)',  inv='12.11.9.7'
          sizes_str='(4-16)',  inv='1.1.1.0 | 1.0.1'
       → zip expanded sizes with the dot-split numbers
    """
    size_qty = {}
    if not inv_detail:
        return size_qty

    inv_clean = inv_detail.strip()

    # ── Format 1: 'size:qty' pipe-separated ──
    if re.search(r'\d+:\d+', inv_clean):
        for pair in re.split(r'[|,;]', inv_clean):
            pair = pair.strip()
            if ':' in pair:
                parts = pair.split(':', 1)
                try:
                    size_qty[parts[0].strip()] = int(parts[1].strip())
                except (ValueError, IndexError):
                    pass
        return size_qty

    # ── Format 2: 'qty/size' pairs ──
    qty_size_pairs = re.findall(r'(\d+)/(\d+)', inv_clean)
    if qty_size_pairs:
        for qty_s, size_s in qty_size_pairs:
            size_qty[size_s] = int(qty_s)
        return size_qty

    # ── Format 3: dot-separated quantities matched against sizes_str ──
    sizes_list = _expand_sizes(sizes_str)
    if sizes_list:
        # Flatten any pipe separators then split on dots/commas
        flat = inv_clean.replace('|', '.').replace(' ', '')
        qty_values = [int(x) for x in re.split(r'[.,]+', flat) if x.isdigit()]
        for sz, qty in zip(sizes_list, qty_values):
            size_qty[sz] = qty

    return size_qty


def _write_excel(results: list[dict], brand_name: str, style_order: list[str] = None) -> str:
    """
    Write results to Excel in flat layout:
      Style No | Color | 0 | 2 | 4 | 6 | 8 | 10 | 12 | 14 | 16 |
      Qty | Unit Cost | Total Cost | List At | MSRP |
      Product Link | Image Link
    One row per style+color. Extra product/image links stack in additional
    rows below (all metadata columns blank on those rows).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] openpyxl not installed — falling back to CSV")
        return _write_csv_fallback(results, brand_name)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = brand_name.replace(" ", "_").replace("&", "and")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # ── Fixed size columns ──
    SIZES = ["0", "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24"]

    # ── Column definitions ──
    # Col index (1-based): name
    HEADERS = (
        ["Style No", "Color"]
        + SIZES
        + ["Qty", "Unit Cost", "Total Cost", "List At", "MSRP",
           "Product Link", "Image Link"]
    )

    # Column letter helpers
    import openpyxl.utils as xl_utils
    def col_letter(idx):  # 1-based
        return xl_utils.get_column_letter(idx)

    # Fixed column positions (1-based)
    C_STYLE     = 1
    C_COLOR     = 2
    C_SIZE_0    = 3                          # first size col
    C_SIZE_LAST = C_SIZE_0 + len(SIZES) - 1 # last size col
    C_QTY       = C_SIZE_LAST + 1
    C_UNIT_COST = C_QTY + 1
    C_TOT_COST  = C_UNIT_COST + 1
    C_LIST_AT   = C_TOT_COST + 1
    C_MSRP      = C_LIST_AT + 1
    C_URL       = C_MSRP + 1
    C_IMG       = C_URL + 1
    TOTAL_COLS  = C_IMG

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = brand_name[:31]

    from openpyxl.styles.borders import Border, Side
    _thin  = Side(style="thin",   color="AAAAAA")
    _thick = Side(style="medium", color="888888")
    _border_main  = Border(top=_thick, bottom=_thin, left=_thin, right=_thin)
    _border_extra = Border(top=_thin,  bottom=_thin, left=_thin, right=_thin)
    _border_blank = Border(top=_thin,  bottom=_thin, left=_thin, right=_thin)

    def apply_border(row, border):
        for c in range(1, TOTAL_COLS + 1):
            ws.cell(row=row, column=c).border = border

    # ── Styles ──
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align   = Alignment(horizontal="center", vertical="center")
    main_fill   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    extra_fill  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    style_font  = Font(bold=True, size=11)
    link_font   = Font(color="0563C1", underline="single")
    num_align   = Alignment(horizontal="center")

    # ── Header row ──
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    # ── Column widths ──
    ws.column_dimensions[col_letter(C_STYLE)].width = 16
    ws.column_dimensions[col_letter(C_COLOR)].width = 18
    for i in range(len(SIZES)):
        ws.column_dimensions[col_letter(C_SIZE_0 + i)].width = 6
    ws.column_dimensions[col_letter(C_QTY)].width       = 8
    ws.column_dimensions[col_letter(C_UNIT_COST)].width  = 12
    ws.column_dimensions[col_letter(C_TOT_COST)].width   = 12
    ws.column_dimensions[col_letter(C_LIST_AT)].width    = 12
    ws.column_dimensions[col_letter(C_MSRP)].width       = 12
    ws.column_dimensions[col_letter(C_URL)].width        = 70
    ws.column_dimensions[col_letter(C_IMG)].width        = 70

    ws.freeze_panes = "A2"

    # ── Write Flat Layout ──
    row_num = 2
    
    last_was_main = False
    
    grand_total_qty = 0
    grand_total_cost = 0.0
    
    for r in results:
        is_secondary = r.get("is_secondary", False)

        if is_secondary:
            # Secondary row: just URL and Image, extra fill, no metadata
            for c in range(1, TOTAL_COLS + 1):
                ws.cell(row=row_num, column=c).fill = extra_fill
            apply_border(row_num, _border_extra)
            url = r.get("Product URL", "")
            img = r.get("Image Link", "")
            url_cell = ws.cell(row=row_num, column=C_URL, value=url)
            if url:
                url_cell.font = link_font
                url_cell.hyperlink = url
            ws.cell(row=row_num, column=C_IMG, value=img)
            row_num += 1
            last_was_main = False
        else:
            # Main row(s)
            
            # Add a separator blank row if the previous row was also output
            if row_num > 2:
                row_num += 1
                
            style_no = r.get("Style No", "")
            

            # ── Parse per-size inventory ──
            inv_detail = r.get("Inventory Detail", "")
            sizes_str  = r.get("Sizes", "")
            size_qty   = _parse_inv_detail(inv_detail, sizes_str)

            # ── Prices ──
            unit_cost_raw = r.get("Wholesale Price") or r.get("Retail Price") or r.get("Price (USD)") or ""
            try:
                unit_cost_val = float(str(unit_cost_raw).replace("$", "").replace(",", "").strip())
            except (ValueError, TypeError):
                unit_cost_val = None

            total_qty_raw = r.get("Total Qty", "")
            try:
                total_qty_val = int(str(total_qty_raw).strip())
            except (ValueError, TypeError):
                total_qty_val = None

            total_cost_val = None
            if unit_cost_val is not None and total_qty_val is not None:
                total_cost_val = round(unit_cost_val * total_qty_val, 2)

            if total_qty_val is not None:
                grand_total_qty += total_qty_val
            if total_cost_val is not None:
                grand_total_cost += total_cost_val

            color_val = r.get("Colors") or r.get("Color") or ""
            msrp_raw = r.get("Retail Price") or ""

            # ROW 1: STYLE NO & SIZES
            ws.cell(row=row_num, column=C_STYLE, value=style_no).font = style_font
            
            for i, sz in enumerate(SIZES):
                qty_for_sz = size_qty.get(sz)
                if qty_for_sz is not None and qty_for_sz != "":
                    # Write the size label in the corresponding column
                    cell = ws.cell(row=row_num, column=C_SIZE_0 + i, value=f"{sz}")
                    cell.alignment = num_align
                    cell.font = style_font

            # Apply fill and border for ROW 1
            for c in range(1, TOTAL_COLS + 1):
                ws.cell(row=row_num, column=c).fill = main_fill
            apply_border(row_num, _border_main)
            
            row_num += 1

            # ROW 2: COLOR & QUANTITIES & OTHER METADATA
            ws.cell(row=row_num, column=C_STYLE, value=color_val) # Color goes in Style column
            
            for i, sz in enumerate(SIZES):
                qty_for_sz = size_qty.get(sz)
                cell = ws.cell(row=row_num, column=C_SIZE_0 + i,
                               value=qty_for_sz if qty_for_sz is not None else "")
                cell.alignment = num_align

            ws.cell(row=row_num, column=C_QTY,
                    value=total_qty_val if total_qty_val is not None else total_qty_raw)
            ws.cell(row=row_num, column=C_UNIT_COST,
                    value=f"${unit_cost_val:.2f}" if unit_cost_val is not None else unit_cost_raw)
            ws.cell(row=row_num, column=C_TOT_COST,
                    value=f"${total_cost_val:.2f}" if total_cost_val is not None else "")
            ws.cell(row=row_num, column=C_LIST_AT, value="")
            ws.cell(row=row_num, column=C_MSRP, value=msrp_raw)

            # URL and Image
            url = r.get("Product URL") or r.get("Product Link") or r.get("URL") or ""
            img = r.get("Image Link") or r.get("Image") or ""
            url_cell = ws.cell(row=row_num, column=C_URL, value=url)
            if url:
                url_cell.font = link_font
                url_cell.hyperlink = url
            ws.cell(row=row_num, column=C_IMG, value=img)

            # Apply fill and border for ROW 2
            for c in range(1, TOTAL_COLS + 1):
                ws.cell(row=row_num, column=c).fill = main_fill
            apply_border(row_num, _border_main)
            
            row_num += 1
            last_was_main = True

    # ── Write GRAND TOTAL Row ──
    row_num += 1 # Add one blank row before total
    ws.cell(row=row_num, column=C_STYLE, value="TOTAL").font = style_font
    ws.cell(row=row_num, column=C_QTY, value=grand_total_qty).font = style_font
    ws.cell(row=row_num, column=C_TOT_COST, value=f"${grand_total_cost:.2f}").font = style_font

    wb.save(filepath)
    print(f"  [OK] Excel saved: {filepath}")
    return filepath

CSV_COLUMNS = [
    "Source PDF", "Style No", "Description", "Color", "Sizes Available",
    "Total Qty", "Inventory Detail", "Wholesale Price", "Retail Price",
    "Image Link", "Season", "Fabric", "Country of Origin", "Code", "Vendor"
]

def _write_csv(results: list[dict], brand_name: str) -> str:
    """Write results to CSV with specific columns for the user."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = brand_name.replace(" ", "_").replace("&", "and")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{ts}.csv"
    filepath = os.path.join(OUTPUT_DIR, filename)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for r in results:
            row = {
                "Source PDF":        r.get("Source PDF", ""),
                "Style No":          r.get("Style No", ""),
                "Description":       r.get("Description", ""),
                "Color":             r.get("Colors") or r.get("Color", ""),
                "Sizes Available":   r.get("Sizes")  or r.get("Sizes Available", ""),
                "Total Qty":         r.get("Total Qty", ""),
                "Inventory Detail":  r.get("Inventory Detail", ""),
                "Wholesale Price":   r.get("Wholesale Price") or r.get("Price (USD)") or "",
                "Retail Price":      r.get("Retail Price", ""),
                "Image Link":        r.get("Image Link", ""),
                "Season":            r.get("Season", ""),
                "Fabric":            r.get("Fabric", ""),
                "Country of Origin": r.get("Country of Origin", ""),
                "Code":              r.get("Code", ""),
                "Vendor":            r.get("Vendor", ""),
            }
            writer.writerow(row)

    print(f"  [OK] CSV saved: {filepath}")
    return filepath


def _write_csv_fallback(results: list[dict], brand_name: str) -> str:
    return _write_csv(results, brand_name)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Product Scraper v4")
    parser.add_argument("--input", "-i", required=True, help="Path to PDF or CSV file")
    parser.add_argument("--brand", "-b", required=True, help="Brand/vendor name")
    parser.add_argument("--force", "-f", action="store_true", help="Force re-scrape")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"[ERROR] Input file not found: {args.input}")
        sys.exit(1)

    run(args.input, args.brand, args.force)
