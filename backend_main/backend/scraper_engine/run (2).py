"""
Product Scraper — Main Runner (v4)
Now returns an Excel file path + updates a progress dict for the API.

Usage (CLI):
    python run.py --input <pdf_or_csv> --brand <brand_name>
"""
import argparse
import csv
import os
import sys
from datetime import datetime

from config import CACHE_COLUMNS, OUTPUT_DIR
from cache import get_cached_styles, get_cached_results_for_vendor, append_to_cache
from pdf_extractor import extract_styles_from_pdf
from csv_reader import read_styles_from_csv
from serper_search import batch_search
from filter import filter_results


def run(input_path: str, brand_name: str, force: bool = False, progress: dict = None):
    """
    Main scraper entry point.

    Args:
        input_path: Path to PDF or CSV file
        brand_name: Vendor/brand name
        force: If True, re-scrape even if style is in cache
        progress: Optional dict updated in-place for API progress tracking

    Returns:
        str: Path to the output Excel file
    """
    if progress is None:
        progress = {}

    def update(step, detail="", pct=0):
        progress["step"] = step
        progress["detail"] = detail
        progress["percent"] = pct
        print(f"  [{pct}%] {step}: {detail}")

    print(f"\n{'='*60}")
    print(f"  Product Scraper — {brand_name}")
    print(f"  Input: {os.path.basename(input_path)}")
    print(f"{'='*60}\n")

    # ── Step 1: Extract style numbers ──
    update("extracting", "Reading input file...", 5)
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".pdf":
        styles = extract_styles_from_pdf(input_path)
    elif ext in (".csv", ".tsv", ".xlsx"):
        styles = read_styles_from_csv(input_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    if not styles:
        raise ValueError("No style numbers found in input file!")

    all_style_numbers = {s["Style No"] for s in styles}
    update("extracting", f"Found {len(all_style_numbers)} unique styles", 10)
    progress["total_styles"] = len(all_style_numbers)

    # ── Step 2: Check cache ──
    update("cache_check", "Checking cache...", 15)
    cached_styles = get_cached_styles(vendor_name=brand_name)
    cached_results = get_cached_results_for_vendor(brand_name)

    already_done = all_style_numbers & cached_styles
    needs_search = all_style_numbers - cached_styles

    if force:
        needs_search = all_style_numbers
        already_done = set()

    progress["cached"] = len(already_done)
    progress["needs_search"] = len(needs_search)
    update("cache_check", f"{len(already_done)} cached, {len(needs_search)} new", 20)

    # ── Step 3: Serper search (parallel — up to 20 concurrent) ──
    new_results = []
    if needs_search:
        styles_to_search = [s for s in styles if s["Style No"] in needs_search]
        total_to_search = len(styles_to_search)

        def _search_progress(completed, total, style_no):
            pct = 20 + int((completed / total) * 55)  # 20% → 75%
            update("searching", f"[{completed}/{total}] {style_no}", pct)

        update("searching", f"Running parallel Serper for {total_to_search} styles (20 workers)...", 25)
        raw_results = batch_search(styles_to_search, brand_name, progress_callback=_search_progress)

        update("filtering", "Filtering results...", 75)
        new_results = filter_results(raw_results, brand_name)

        # Save to cache
        if new_results:
            cache_rows = []
            for r in new_results:
                cache_rows.append({
                    "Style No": r.get("Style No", ""),
                    "Total Qty": r.get("Total Qty", ""),
                    "Product URL": r.get("Product URL", ""),
                    "Price (USD)": r.get("Price (USD)", ""),
                    "Image Link": r.get("Image Link", ""),
                    "Vendor Name": brand_name,
                })
            append_to_cache(cache_rows)
    else:
        update("searching", "All styles in cache — no Serper calls needed", 75)

    # ── Step 4: Combine results ──
    update("building_output", "Building Excel...", 80)
    combined = []
    styles_in_output = set()

    for r in cached_results:
        if r.get("Style No") in all_style_numbers:
            combined.append(r)
            styles_in_output.add(r.get("Style No"))

    for style_no in all_style_numbers:
        if style_no in cached_styles and style_no not in styles_in_output:
            combined.append({
                "Style No": style_no,
                "Total Qty": next((s.get("Total Qty") for s in styles if s["Style No"] == style_no), ""),
                "Product URL": "",
                "Price (USD)": "",
                "Image Link": "",
                "Vendor Name": brand_name,
            })
            styles_in_output.add(style_no)

    for r in new_results:
        combined.append({
            "Style No": r.get("Style No", ""),
            "Total Qty": r.get("Total Qty", ""),
            "Product URL": r.get("Product URL", ""),
            "Price (USD)": r.get("Price (USD)", ""),
            "Image Link": r.get("Image Link", ""),
            "Vendor Name": brand_name,
        })

    # ── Step 4b: Deduplicate ──
    update("deduplicating", "Removing duplicate URLs...", 85)
    seen_keys = set()
    deduped = []
    for r in combined:
        url = (r.get("Product URL") or "").strip().lower().rstrip("/")
        if "?" in url:
            url = url.split("?")[0]
        style = (r.get("Style No") or "").strip()
        key = (style, url) if url else (style, id(r))  # keep no-URL rows
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(r)
    removed = len(combined) - len(deduped)
    if removed:
        update("deduplicating", f"Removed {removed} duplicate URLs", 87)
    combined = deduped

    # ── Step 5: Write Excel ──
    update("writing_excel", "Saving Excel file...", 90)
    # Pass original style order so Excel preserves PDF/CSV sequence
    original_order = [s["Style No"] for s in styles]
    excel_path = _write_excel(combined, brand_name, style_order=original_order)
    progress["output_file"] = excel_path

    # ── Summary ──
    unique_with_urls = len({r.get("Style No") for r in combined if r.get("Product URL")})
    total_listings = len([r for r in combined if r.get("Product URL")])

    progress["total_listings"] = total_listings
    progress["styles_with_results"] = unique_with_urls

    update("done", f"{total_listings} listings for {unique_with_urls} styles", 100)

    print(f"\n{'='*60}")
    print(f"  ✅ DONE — {brand_name}")
    print(f"  Input styles:        {len(all_style_numbers)}")
    print(f"  From cache:          {len(already_done)}")
    print(f"  New Serper calls:    {len(needs_search)}")
    print(f"  Total listings:      {total_listings}")
    print(f"  Styles with results: {unique_with_urls}")
    print(f"  Excel: {excel_path}")
    print(f"{'='*60}\n")

    return excel_path


def _write_excel(results: list[dict], brand_name: str, style_order: list[str] = None) -> str:
    """Write results to an Excel file — 2 or 3 columns depending on qty data.
    3 empty rows between each style group. Preserves original PDF/CSV order."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠️ openpyxl not installed — falling back to CSV")
        return _write_csv_fallback(results, brand_name)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = brand_name.replace(" ", "_").replace("&", "and")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = brand_name[:31]

    # ── Decide if we include Total Qty column ──
    # Only add it when at least one style has qty data (typically from PDFs)
    from gsheet_writer import group_by_style
    grouped = group_by_style(results)

    has_qty = any(
        data.get("Total Qty") not in (None, "", 0)
        for data in grouped.values()
    )

    if not has_qty:
        # also check raw results
        has_qty = any(
            r.get("Total Qty") not in (None, "", 0)
            for r in results
        )

    headers = ["Style No", "Product URL"]
    col_widths = {"A": 18, "B": 80}
    url_col = 2

    if has_qty:
        headers = ["Style No", "Total Qty", "Product URL"]
        col_widths = {"A": 18, "B": 12, "C": 80}
        url_col = 3

    # ── Header ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    style_font = Font(bold=True, size=11)
    style_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    # Preserve original order from PDF/CSV; fall back to alphabetical
    if style_order:
        seen_order = set()
        sorted_styles = []
        for s in style_order:
            if s in grouped and s not in seen_order:
                sorted_styles.append(s)
                seen_order.add(s)
        # Append any styles not in the original order (e.g. from cache)
        for s in grouped:
            if s not in seen_order:
                sorted_styles.append(s)
    else:
        sorted_styles = sorted(grouped.keys())

    # Build qty lookup from input styles
    qty_lookup = {}
    for r in results:
        sn = r.get("Style No", "")
        qty = r.get("Total Qty")
        if sn and qty not in (None, "", 0):
            qty_lookup.setdefault(sn, qty)

    for idx, style_no in enumerate(sorted_styles):
        data = grouped[style_no]
        listings = data["listings"]

        # Style row — fill background on all columns
        style_cell = ws.cell(row=row_num, column=1, value=style_no)
        style_cell.font = style_font
        style_cell.fill = style_fill
        for c in range(2, len(headers) + 1):
            ws.cell(row=row_num, column=c).fill = style_fill

        # Total Qty on style row (only if column exists)
        if has_qty:
            qty_val = qty_lookup.get(style_no)
            if qty_val not in (None, "", 0):
                qty_cell = ws.cell(row=row_num, column=2, value=qty_val)
                qty_cell.font = style_font
                qty_cell.fill = style_fill
                qty_cell.alignment = Alignment(horizontal="center")

        if listings:
            # First URL on same row as style
            url_cell = ws.cell(row=row_num, column=url_col, value=listings[0].get("Product URL", ""))
            url_cell.font = link_font
            url_cell.fill = style_fill
            if listings[0].get("Product URL"):
                url_cell.hyperlink = listings[0]["Product URL"]
            row_num += 1

            # Remaining URLs
            for listing in listings[1:]:
                url_cell = ws.cell(row=row_num, column=url_col, value=listing.get("Product URL", ""))
                url_cell.font = link_font
                if listing.get("Product URL"):
                    url_cell.hyperlink = listing["Product URL"]
                row_num += 1
        else:
            # Style with 0 links — still gets its row
            row_num += 1

        # 3 empty rows between style groups
        if idx < len(sorted_styles) - 1:
            row_num += 3

    # ── Column widths ──
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze header ──
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"  ✅ Excel saved: {filepath}")
    return filepath


def _write_csv_fallback(results: list[dict], brand_name: str) -> str:
    """Fallback if openpyxl is not installed."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_name = brand_name.replace(" ", "_").replace("&", "and")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{ts}.csv"
    filepath = os.path.join(OUTPUT_DIR, filename)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CACHE_COLUMNS)
        writer.writeheader()
        for r in results:
            writer.writerow({col: r.get(col, "") for col in CACHE_COLUMNS})

    print(f"  ✅ CSV saved: {filepath}")
    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Product Scraper v4")
    parser.add_argument("--input", "-i", required=True, help="Path to PDF or CSV file")
    parser.add_argument("--brand", "-b", required=True, help="Brand/vendor name")
    parser.add_argument("--force", "-f", action="store_true", help="Force re-scrape")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ Input file not found: {args.input}")
        sys.exit(1)

    run(args.input, args.brand, args.force)
